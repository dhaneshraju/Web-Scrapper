from django.shortcuts import render
from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import wget
import time
import requests
import openpyxl
import random
import os

# Create your views here.
a = random.randint(0, 101)
ran = str(a)

def home(request):
    return render(request, 'base.html')

def jobs(request):
    return render(request, 'jobs.html')

def movie(request):
    return render(request, 'movies.html')

def insta(request):
    return render(request, 'insta.html')

def insta_hashtag(request):
    return render(request, 'insta_hashtag.html')

def insta_user(request):
    return render(request, 'insta_user.html')

def ads(request):
    return render(request, 'ads.html')

def linkedin(request):
    return render(request, 'linkedin.html')

def profile(request):
    return render(request, 'linkedin_profile.html')

def company(request):
    return render(request, 'linkedin_company.html')

def demo_linkedin(request):
    return render(request, 'demo.html')


def search(request):
    try:
        topic = request.GET.get('desigination')
        pre_location = request.GET.get('prefer_location')
        exp = request.GET.get('exp')
        date_posted = request.GET.get('date_of_filters')
        excel = openpyxl.Workbook()
        sheet = excel.active
        sheet.title = "Job List"
        sheet.append(['Position', 'Company Name', 'Location', 'Description', 'Skills', 'More Info'])
        html_text = requests.get(f'https://www.timesjobs.com/candidate/job-search.html?searchType=personalizedSearch&from=submit&txtKeywords={topic}&txtLocation={pre_location}&cboWorkExp1={exp}').text

        soup = BeautifulSoup(html_text, 'lxml')
        jobs = soup.find_all('li', class_='clearfix job-bx wht-shd-bx')
        try:
            for job in jobs:
                published_data = job.find('span', class_='sim-posted').span.text
                if f'{date_posted}' in published_data:
                    try:
                        title = job.header.h2.a.text
                    except:
                        title = job.find('strong', class_='blkclor').text
                    company_name = job.find('h3', class_='joblist-comp-name').text.replace(' ', '')
                    location = job.find('span', title_="").text
                    position = job.find('ul', 'list-job-dtl clearfix').li.text
                    if 'Job Description:' in position:
                        position = position.replace('Job Description:', '')
                    skills = job.find('span', 'srp-skills').text.replace(' ', '')
                    more_info = job.header.h2.a['href']
                    sheet.append([title.strip(), company_name.strip(), location.strip(), position.strip(), skills.strip(),
                                  more_info])
                    excel.save(f'Jobs{ran}.xlsx')
                    print('Process Completed, Look the Directory')


        except Exception as e:
            print('Error occurs, Please try again...')

        return render(request, 'search.html')
    except Exception as e:
        return render(request, 'error.html')



def trendingmovies(request):
    try:
        excel = openpyxl.Workbook()
        sheet = excel.active
        sheet.title = "Trending-Movies"
        sheet.append(['Rank', 'Movie Name', 'Year', 'Rating'])

        html_text = requests.get('https://www.imdb.com/chart/top/').text
        try:
            soup = BeautifulSoup(html_text, 'lxml')
            html_content = soup.find('tbody', class_='lister-list').find_all('tr')
            for html_contents in html_content:
                rank = html_contents.find('td', class_='titleColumn').get_text(strip=True).split('.')[0]
                movie_name = html_contents.find('td', class_='titleColumn').a.text
                year = html_contents.find('span', class_='secondaryInfo').text.replace('(', '').replace(')', '')
                rating = html_contents.find('td', class_='ratingColumn imdbRating').strong.text
                # print(rank, movie_name, year, rating)
                sheet.append([rank, movie_name, year, rating])
                excel.save(f'trendingmovies{ran}.xlsx')
            print('Process Completed, Look the Directory')
        except Exception as e:
            print('Error Occurs,  Please try again...')
        return render(request, 'search.html')

    except Exception as e:
        return render(request, 'error.html')

def genres(request):
    try:
        excel = openpyxl.Workbook()
        sheet = excel.active
        sheet.title = "Gener-Movies"
        sheet.append(['Rank', 'Movie Name', 'Genre', 'Year', 'Rating'])

        need_genre = request.GET.get('desigination')

        html_text = requests.get(f'https://www.imdb.com/search/title/?genres={need_genre}&explore=').text

        soup = BeautifulSoup(html_text, 'lxml')
        try:
            source = soup.find('div', class_='lister list detail sub-list')
            temp = source.find('div', class_='lister-list')
            html_content = temp.find_all('div', class_='lister-item mode-advanced')
            for html_contents in html_content:
                rank = html_contents.find('div', class_='lister-item-content').h3.span.text.replace('.', '')
                movie_name = html_contents.find('div', class_='lister-item-content').h3.a.text
                genre = html_contents.find('span', class_='genre').text.strip()
                year = html_contents.find('span', class_='lister-item-year text-muted unbold').text.replace('(',
                                                                                                            '').replace(')',
                                                                                                                        '')
                try:
                    rating = html_contents.find('div', class_='inline-block ratings-imdb-rating').strong.text
                except:
                    rating = 'N/A'
                # print(rank, movie_name, genre, year, rating)
                sheet.append([rank, movie_name, genre, year, rating])
                excel.save(f'{need_genre}movies{ran}.xlsx')
            print('Process Completed, Look the Directory')
        except:
            print('Error Occurs')
        return render(request, 'search.html')

    except Exception as e:
        return render(request, 'error.html')


def populars(request):
    try:
        excel = openpyxl.Workbook()
        sheet = excel.active
        sheet.title = "Popular-Movies"
        sheet.append(['Rank', 'Movie Name', 'Year', 'Rating'])

        html_text = requests.get('https://www.imdb.com/chart/moviemeter/').text
        try:
            soup = BeautifulSoup(html_text, 'lxml')
            html_content = soup.find('tbody', class_='lister-list').find_all('tr')
            for html_contents in html_content:
                rank = html_contents.find('div', class_='velocity').text.split()[0]
                movie_name = html_contents.find('td', class_='titleColumn').a.text
                year = html_contents.find('span', class_='secondaryInfo').text.replace('(', '').replace(')', '')
                try:
                    rating = html_contents.find('strong', title_='').text
                except:
                    rating = 'N/A'
                # print(rank, movie_name, year, rating)
                sheet.append([rank, movie_name, year, rating])
                excel.save(f'popularmovies{ran}.xlsx')
            print('Process Completed, Look the Directory')
        except:
            print('Error Occurs')
        return render(request, 'search.html')

    except Exception as e:
        return render(request, 'error.html')

def target_hashtag(request):
    try:
        user = request.GET.get('user')
        pwd = request.GET.get('pass')
        page_ig = request.GET.get('hashtag')

        options = webdriver.ChromeOptions()
        options.add_argument("--lang=en")
        browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        browser.get("https://www.instagram.com/login")
        time.sleep(5)
        username_input = browser.find_element(By.CSS_SELECTOR, "input[name='username']")
        password_input = browser.find_element(By.CSS_SELECTOR, "input[name='password']")
        username_input.send_keys(user)
        password_input.send_keys(pwd)
        time.sleep(5)
        login_button = browser.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()
        time.sleep(2.5)
        browser.get(f"https://www.instagram.com/explore/tags/{page_ig}/")
        time.sleep(5)
        try:
            browser.execute_script("window.scrollTo(0,4000);")
            images = browser.find_elements(By.TAG_NAME, 'img')
            images = [image.get_attribute('src') for image in images]
            path = os.getcwd()
            path = os.path.join(path, page_ig)
            path = path + ran
            os.mkdir(path)
            counter = 0
            for image in images:
                save_as = os.path.join(path, page_ig + f'{counter}' + '.jpg')
                wget.download(image, save_as)
                counter += 1


        except Exception as e:
            print('Error Occurs, Please try again...')
        return render(request, 'search.html')

    except Exception as e:
        return render(request, 'error.html')

def target_user(request):
    try:
        user = request.GET.get('user')
        pwd = request.GET.get('pass')
        page_ig = request.GET.get('target_user')

        options = webdriver.ChromeOptions()
        options.add_argument("--lang=en")
        browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        browser.get("https://www.instagram.com/login")
        time.sleep(5)
        username_input = browser.find_element(By.CSS_SELECTOR, "input[name='username']")
        password_input = browser.find_element(By.CSS_SELECTOR, "input[name='password']")
        username_input.send_keys(user)
        password_input.send_keys(pwd)
        time.sleep(5)
        login_button = browser.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()
        time.sleep(2.5)
        browser.get(f"https://www.instagram.com/{page_ig}")
        time.sleep(5)
        try:
            browser.execute_script("window.scrollTo(0,4000);")
            images = browser.find_elements(By.TAG_NAME, 'img')
            images = [image.get_attribute('src') for image in images]
            path = os.getcwd()
            path = os.path.join(path, page_ig)
            path = path + ran
            os.mkdir(path)
            counter = 0
            for image in images:
                save_as = os.path.join(path, page_ig + f'{counter}' + '.jpg')
                wget.download(image, save_as)
                counter += 1


        except Exception as e:
            print('Error Occurs, Please try again...')
        return render(request, 'search.html')

    except Exception as e:
        return render(request, 'error.html')




def new_search(request):
    try:
        search = request.GET.get('search')
        final_url = f'https://london.craigslist.org/search/sss?query={search}'
        response = requests.get(final_url)
        BASE_IMAGE_URL = 'https://images.craigslist.org/{}_300x300.jpg'
        data = response.text
        soup = BeautifulSoup(data, features='html.parser')

        post_listings = soup.find_all('li', {'class': 'result-row'})

        final_postings = []

        for post in post_listings:
            post_title = post.find(class_='result-title').text
            post_url = post.find('a').get('href')

            if post.find(class_='result-price'):
                post_price = post.find(class_='result-price').text
            else:
                post_price = 'N/A'

            if post.find(class_='result-image').get('data-ids'):
                post_image_id = post.find(class_='result-image').get('data-ids').split(',')[0].split(':')[1]
                post_image_url = BASE_IMAGE_URL.format(post_image_id)
                print(post_image_url)
            else:
                post_image_url = 'https://craigslist.org/images/peace.jpg'

            final_postings.append((post_title, post_url, post_price, post_image_url))

        stuff_for_frontend = {
            'search': search,
            'final_postings': final_postings,
        }

        return render(request, 'ads_search.html', stuff_for_frontend)

    except Exception as e:
        return render(request, 'error.html')


def linkedin_profile(request):
    try:
        driver_service = Service('D:/publish/chromedriver.exe')

        browser = webdriver.Chrome(service=driver_service)
        browser.get('https://www.linkedin.com/uas/login')

        user = request.GET.get('user')
        pwd = request.GET.get('pass')
        people = request.GET.get('user_link')

        username_input = browser.find_element(By.CSS_SELECTOR, "input[name='session_key']")
        password_input = browser.find_element(By.CSS_SELECTOR, "input[name='session_password']")
        username_input.send_keys(user)
        password_input.send_keys(pwd)
        time.sleep(5)

        login_button = browser.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()
        time.sleep(5)

        profile_link = f'{people}'
        browser.get(profile_link)

        src = browser.page_source
        soup = BeautifulSoup(src, 'lxml')

        html_text = soup.find('div', class_='ph5')
        html_text1 = soup.find('div', class_='display-flex ph5 pv3')
        name = html_text.find('h1', class_='text-heading-xlarge inline t-24 v-align-middle break-words').text
        whats_doing = html_text.find('div', class_='text-body-medium break-words').text.strip()
        whats_doing = str(whats_doing)
        education = html_text.find('div',
                                   class_='inline-show-more-text inline-show-more-text--is-collapsed inline-show-more-text--is-collapsed-with-line-clamp inline').text.strip()
        location = html_text.find('span', class_='text-body-small inline t-black--light break-words').text.strip()
        about = html_text1.find('span').text.strip()
        with open(f'{name}.txt', 'w') as f:
            try:
                f.write(f'Name: {name} \n')
            except:
                name = 'N/A'
                f.write(f'Name: {name} \n')
            try:
                f.write(f'Working: {whats_doing} \n')
            except:
                whats_doing = 'N/A'
                f.write(f'Working: {whats_doing} \n')
            try:
                f.write(f'Location: {location} \n')
            except:
                location = 'N/A'
                f.write(f'Location: {location} \n')
            try:
                f.write(f'About: {about} \n')
            except:
                about = 'N/A'
                f.write(f'About: {about} \n')
        print("file saved look directry")

        return render(request, 'search.html')

    except Exception as e:
        return render(request, 'error.html')

def linkedin_company(request):
    try:
        driver_service = Service('D:/publish/chromedriver.exe')

        browser = webdriver.Chrome(service=driver_service)
        browser.get('https://www.linkedin.com/uas/login')

        user = request.GET.get('user')
        pwd = request.GET.get('pass')
        dynamic_company = request.GET.get('company')

        username_input = browser.find_element(By.CSS_SELECTOR, "input[name='session_key']")
        password_input = browser.find_element(By.CSS_SELECTOR, "input[name='session_password']")
        username_input.send_keys(user)
        password_input.send_keys(pwd)
        time.sleep(5)

        login_button = browser.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()
        time.sleep(5)

        profile_link = f'https://www.linkedin.com/company/{dynamic_company}/'
        browser.get(profile_link)

        src = browser.page_source
        soup = BeautifulSoup(src, 'lxml')

        html_text = soup.find('div', class_='ph5 pt3')
        html_text1 = soup.find('div', class_='ph5 pb5')
        company_name = html_text.find('h1', class_="ember-view t-24 t-black t-bold full-width").span.text.strip()
        working = html_text.find('div', class_='org-top-card-summary-info-list__info-item').text.strip()
        location = html_text.find('div', class_='inline-block').div.text.strip()
        about = html_text1.find('div', class_='t-14 t-black--light full-width break-words ember-view').text.strip()

        with open(f'{company_name}.txt', 'w') as f:
            try:
                f.write(f'Company Name: {company_name} \n')
            except:
                company_name = 'N/A'
                f.write(f'Company Name: {company_name} \n')
            try:
                f.write(f'Working: {working} \n')
            except:
                working = 'N/A'
                f.write(f'Working: {working} \n')
            try:
                f.write(f'Location: {location} \n')
            except:
                location = 'N/A'
                f.write(f'Location: {location} \n')
            try:
                f.write(f'About: {about} \n')
            except:
                about = 'N/A'
                f.write(f'About: {about} \n')
        print("file saved look directry")

        return render(request, 'search.html')

    except Exception as e:
        return render(request, 'error.html')


def linkedin_demo(request):
    try:
        driver_service = Service('D:/publish/chromedriver.exe')

        browser = webdriver.Chrome(service=driver_service)
        browser.get('https://www.linkedin.com/uas/login')

        user = request.GET.get('user')
        pwd = request.GET.get('pass')

        list_linkedin = ['https://www.linkedin.com/in/rishimehta/', 'https://www.linkedin.com/in/anandhu-murali/']

        username_input = browser.find_element(By.CSS_SELECTOR, "input[name='session_key']")
        password_input = browser.find_element(By.CSS_SELECTOR, "input[name='session_password']")
        username_input.send_keys(user)
        password_input.send_keys(pwd)
        time.sleep(5)

        login_button = browser.find_element(By.XPATH, "//button[@type='submit']")
        login_button.click()
        time.sleep(5)
        for i in range (len(list_linkedin)):

            profile_link = list_linkedin[i]
            browser.get(profile_link)

            src = browser.page_source
            soup = BeautifulSoup(src, 'lxml')

            html_text = soup.find('div', class_='ph5')
            html_text1 = soup.find('div', class_='display-flex ph5 pv3')
            name = html_text.find('h1', class_='text-heading-xlarge inline t-24 v-align-middle break-words').text
            whats_doing = html_text.find('div', class_='text-body-medium break-words').text.strip()
            location = html_text.find('span', class_='text-body-small inline t-black--light break-words').text.strip()
            about = html_text1.find('span').text.strip()
            with open(f'{name}.txt', 'w') as f:
                try:
                    f.write(f'Name: {name} \n')
                except:
                    name = 'N/A'
                    f.write(f'Name: {name} \n')
                try:
                    f.write(f'Working: {whats_doing} \n')
                except:
                    whats_doing = 'N/A'
                    f.write(f'Working: {whats_doing} \n')
                try:
                    f.write(f'Location: {location} \n')
                except:
                    location = 'N/A'
                    f.write(f'Location: {location} \n')
                try:
                    f.write(f'About: {about} \n')
                except:
                    about = 'N/A'
                    f.write(f'About: {about} \n')
            print("file saved look directry")

        return render(request, 'search.html')


    except Exception as e:
        return render(request, 'error.html')